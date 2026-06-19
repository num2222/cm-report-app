"""
CM Report Export Server (Production & Render Compatible)
วางไฟล์นี้ในโฟลเดอร์เดียวกับ Template_Daily_Report.xlsx และ Template_Month_Report.xlsx
สำหรับรันบน Server เช่น Render ด้วยคำสั่ง: gunicorn cm_server:app
"""
from flask import Flask, request, send_file, jsonify
import json, shutil, io, os
from openpyxl import load_workbook
from pathlib import Path
from datetime import datetime

app = Flask(__name__)

BASE = Path(__file__).parent
TMPL_DAILY  = BASE / 'Template_Daily_Report.xlsx'
TMPL_MONTH  = BASE / 'Template_Month_Report.xlsx'

STD_LABELS = {
    'A': '5-30 นาที', 'B': '1-3 ชม.', 'C': '3 ชม.-1 วัน',
    'D': '1-7 วัน',   'E': '7-14 วัน', 'F': '1 เดือน'
}

def fmt_be(iso_date):
    """2026-06-10 → 10/06/69"""
    if not iso_date: return ''
    p = iso_date.split('-')
    be = int(p[0]) + 543
    return f"{p[2]}/{p[1]}/{str(be)[-2:]}\"\n"

def kpi_sym(v):
    return 'P' if v == 'pass' else ('O' if v == 'fail' else '')

# ── CORS & Preflight Middleware (รองรับการเรียกข้าม Domain บน Render) ──
@app.after_request
def after_request(response):
    response.headers.add('Access-Control-Allow-Origin', '*')
    response.headers.add('Access-Control-Allow-Headers', 'Content-Type,Authorization')
    response.headers.add('Access-Control-Allow-Methods', 'GET,PUT,POST,DELETE,OPTIONS')
    return response

# จัดการ OPTIONS request สำหรับการตรวจสอบก่อนส่งข้อมูล (Preflight CORS)
@app.route('/', defaults={'path': ''}, methods=['OPTIONS'])
@app.route('/<path:path>', methods=['OPTIONS'])
def options_route(path):
    return '', 200

# ── Routes การทำงานของระบบ ──

@app.route('/export/daily', methods=['POST'])
def export_daily():
    data = request.json or {}
    cases = data.get('cases', [])
    date_iso = data.get('date', '') # YYYY-MM-DD
    
    if not TMPL_DAILY.exists():
        return jsonify({'error': 'Template_Daily_Report.xlsx not found'}), 500

    wb = load_workbook(TMPL_DAILY)
    
    # 1. เขียนชีต Daily CM (หน้าสรุป)
    if 'Daily CM' in wb.sheetnames:
        ws_dc = wb['Daily CM']
        ws_dc.cell(row=7, column=3).value = fmt_be(date_iso)
        
        # นับจำนวนงานตามเงื่อนไขเพื่อสรุปหน้าแรก
        t_mtb = [c for c in cases if c.get('zone') == 'MTB']
        t_z2  = [c for c in cases if c.get('zone') == 'Z2']
        t_fz  = [c for c in cases if c.get('zone') == 'FZ']
        t_sat = [c for c in cases if c.get('zone') == 'SAT1']
        
        # เขียนตัวเลขลงในช่องหน้าสรุป (อิงตำแหน่งแถวเดิม)
        # สรุปรวม
        ws_dc.cell(row=13, column=3).value = len(cases)
        ws_dc.cell(row=14, column=3).value = len([c for c in cases if c.get('status') == 'done'])
        ws_dc.cell(row=15, column=3).value = len([c for c in cases if c.get('status') == 'pending'])
        
        # MTB
        ws_dc.cell(row=20, column=4).value = len(t_mtb)
        ws_dc.cell(row=21, column=4).value = len([c for c in t_mtb if c.get('status') == 'done'])
        ws_dc.cell(row=22, column=4).value = len([c for c in t_mtb if c.get('status') == 'pending'])
        # Z2
        ws_dc.cell(row=24, column=4).value = len(t_z2)
        ws_dc.cell(row=25, column=4).value = len([c for c in t_z2 if c.get('status') == 'done'])
        ws_dc.cell(row=26, column=4).value = len([c for c in t_z2 if c.get('status') == 'pending'])
        # FZ
        ws_dc.cell(row=28, column=4).value = len(t_fz)
        ws_dc.cell(row=29, column=4).value = len([c for c in t_fz if c.get('status') == 'done'])
        ws_dc.cell(row=30, column=4).value = len([c for c in t_fz if c.get('status') == 'pending'])
        # SAT1
        ws_dc.cell(row=32, column=4).value = len(t_sat)
        ws_dc.cell(row=33, column=4).value = len([c for c in t_sat if c.get('status') == 'done'])
        ws_dc.cell(row=34, column=4).value = len([c for c in t_sat if c.get('status') == 'pending'])

        # สรุปผล KPI ส่วนที่ 1 และ 2 หน้าแรก
        k1_pass = len([c for c in cases if c.get('kpi1') == 'pass'])
        k1_fail = len([c for c in cases if c.get('kpi1') == 'fail'])
        k1_total = k1_pass + k1_fail
        ws_dc.cell(row=39, column=4).value = k1_pass
        ws_dc.cell(row=39, column=6).value = round((k1_pass / k1_total * 100), 1) if k1_total > 0 else 0
        ws_dc.cell(row=41, column=4).value = k1_fail
        ws_dc.cell(row=41, column=6).value = round((k1_fail / k1_total * 100), 1) if k1_total > 0 else 0

        k2_pass = len([c for c in cases if c.get('kpi2') == 'pass'])
        k2_fail = len([c for c in cases if c.get('kpi2') == 'fail'])
        k2_total = k2_pass + k2_fail
        ws_dc.cell(row=39, column=10).value = k2_pass
        ws_dc.cell(row=39, column=12).value = round((k2_pass / k2_total * 100), 1) if k2_total > 0 else 0
        ws_dc.cell(row=41, column=10).value = k2_fail
        ws_dc.cell(row=41, column=12).value = round((k2_fail / k2_total * 100), 1) if k2_total > 0 else 0

    # 2. เขียนแยกรายชีตตาม Zone (MTB, Z2, FZ, SAT1)
    for zname in ['MTB', 'Z2', 'FZ', 'SAT1']:
        if zname in wb.sheetnames:
            ws = wb[zname]
            ws.cell(row=4, column=5).value = fmt_be(date_iso) # ใส่วันที่ประจำวันของโซนนั้นๆ
            
            z_cases = [c for c in cases if c.get('zone') == zname]
            row_idx = 7
            for idx, c in enumerate(z_cases, start=1):
                ws.cell(row=row_idx, column=1).value = idx
                ws.cell(row=row_idx, column=2).value = c.get('jobNo', '')
                ws.cell(row=row_idx, column=3).value = c.get('sapNo', '')
                ws.cell(row=row_idx, column=4).value = c.get('timeReq', '')
                ws.cell(row=row_idx, column=5).value = c.get('location', '')
                ws.cell(row=row_idx, column=6).value = c.get('detail', '')
                
                # ฝั่งรับแจ้ง (KPI 1)
                ws.cell(row=row_idx, column=7).value  = c.get('kpi1_no', '')
                ws.cell(row=row_idx, column=8).value  = STD_LABELS.get(c.get('kpi1_no',''), '')
                ws.cell(row=row_idx, column=9).value  = c.get('timeApprove', '')
                ws.cell(row=row_idx, column=10).value = c.get('timeArrive', '')
                ws.cell(row=row_idx, column=11).value = kpi_sym(c.get('kpi1', ''))
                
                # ฝั่งดำเนินการ (KPI 2)
                ws.cell(row=row_idx, column=12).value = c.get('actionDetail', '')
                ws.cell(row=row_idx, column=13).value = c.get('kpi2_no', '')
                ws.cell(row=row_idx, column=14).value = STD_LABELS.get(c.get('kpi2_no',''), '')
                ws.cell(row=row_idx, column=15).value = c.get('timeFinish', '')
                ws.cell(row=row_idx, column=16).value = kpi_sym(c.get('kpi2', ''))
                
                ws.cell(row=row_idx, column=17).value = c.get('remark', '')
                row_idx += 1

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    fname = f"CM_Daily_{date_iso}.xlsx"
    return send_file(out, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=fname)


@app.route('/export/monthly', methods=['POST'])
def export_monthly():
    data = request.json or {}
    cases = data.get('cases', [])
    month_iso = data.get('month', '') # YYYY-MM
    
    if not TMPL_MONTH.exists():
        return jsonify({'error': 'Template_Month_Report.xlsx not found'}), 500

    wb = load_workbook(TMPL_MONTH)
    
    # วิ่งไล่เติมชีตของแต่ละโซนในรายงานรายเดือน
    for zname in ['MTB', 'Z2', 'FZ', 'SAT1']:
        if zname in wb.sheetnames:
            ws = wb[zname]
            z_cases = [c for c in cases if c.get('zone') == zname]
            row_idx = 3 # รายเดือนเริ่มเขียนแถว 3
            for idx, c in enumerate(z_cases, start=1):
                ws.cell(row=row_idx, column=1).value = idx
                ws.cell(row=row_idx, column=2).value = c.get('jobNo', '')
                ws.cell(row=row_idx, column=3).value = c.get('sapNo', '')
                ws.cell(row=row_idx, column=4).value = c.get('timeReq', '')
                ws.cell(row=row_idx, column=5).value = c.get('location', '')
                ws.cell(row=row_idx, column=6).value = c.get('detail', '')
                
                ws.cell(row=row_idx, column=7).value  = c.get('kpi1_no', '')
                ws.cell(row=row_idx, column=8).value  = STD_LABELS.get(c.get('kpi1_no',''), '')
                ws.cell(row=row_idx, column=9).value  = c.get('timeApprove', '')
                ws.cell(row=row_idx, column=10).value = c.get('timeArrive', '')
                ws.cell(row=row_idx, column=11).value = kpi_sym(c.get('kpi1', ''))
                
                ws.cell(row=row_idx, column=12).value = c.get('actionDetail', '')
                ws.cell(row=row_idx, column=13).value = c.get('kpi2_no', '')
                ws.cell(row=row_idx, column=14).value = STD_LABELS.get(c.get('kpi2_no',''), '')
                ws.cell(row=row_idx, column=15).value = c.get('timeFinish', '')
                ws.cell(row=row_idx, column=16).value = kpi_sym(c.get('kpi2', ''))
                ws.cell(row=row_idx, column=17).value = kpi_sym(c.get('kpi3', '')) # KPI3 เพิ่มมาในตารางรายเดือน
                ws.cell(row=row_idx, column=18).value = c.get('remark', '')
                row_idx += 1

    # ชีตสรุปรวมรายเดือน Month CM
    if 'Month CM' in wb.sheetnames:
        ws_mc = wb['Month CM']
        
        # แปลงหัวกระดาษเป็น พ.ศ. เช่น 2026-06 -> 06/2569
        if len(month_iso) == 7:
            p = month_iso.split('-')
            ws_mc.cell(row=7, column=3).value = f"{p[1]}/{int(p[0])+543}"

        t_mtb = [c for c in cases if c.get('zone') == 'MTB']
        t_z2  = [c for c in cases if c.get('zone') == 'Z2']
        t_fz  = [c for c in cases if c.get('zone') == 'FZ']
        t_sat = [c for c in cases if c.get('zone') == 'SAT1']

        # เติมตัวเลขสรุปตารางบนหน้าแรกรายเดือน
        ws_mc.cell(row=13, column=3).value = len(cases)
        ws_mc.cell(row=14, column=3).value = len([c for c in cases if c.get('status') == 'done'])
        ws_mc.cell(row=15, column=3).value = len([c for c in cases if c.get('status') == 'pending'])

        # รายโซน
        ws_mc.cell(row=19, column=7).value = len(t_mtb)
        ws_mc.cell(row=20, column=7).value = len([c for c in t_mtb if c.get('status') == 'done'])
        ws_mc.cell(row=21, column=7).value = len([c for c in t_mtb if c.get('status') == 'pending'])

        ws_mc.cell(row=22, column=7).value = len(t_z2)
        ws_mc.cell(row=23, column=7).value = len([c for c in t_z2 if c.get('status') == 'done'])
        ws_mc.cell(row=24, column=7).value = len([c for c in t_z2 if c.get('status') == 'pending'])

        ws_mc.cell(row=25, column=7).value = len(t_fz)
        ws_mc.cell(row=26, column=7).value = len([c for c in t_fz if c.get('status') == 'done'])
        ws_mc.cell(row=27, column=7).value = len([c for c in t_fz if c.get('status') == 'pending'])

        ws_mc.cell(row=28, column=7).value = len(t_sat)
        ws_mc.cell(row=29, column=7).value = len([c for c in t_sat if c.get('status') == 'done'])
        ws_mc.cell(row=30, column=7).value = len([c for c in t_sat if c.get('status') == 'pending'])

        # สรุป KPI1, KPI2, KPI3 รายเดือนตอนท้าย
        k1p = len([c for c in cases if c.get('kpi1') == 'pass'])
        k1f = len([c for c in cases if c.get('kpi1') == 'fail'])
        k1t = k1p + k1f or 1

        k2p = len([c for c in cases if c.get('kpi2') == 'pass'])
        k2f = len([c for c in cases if c.get('kpi2') == 'fail'])
        k2t = k2p + k2f or 1

        k3p = len([c for c in cases if c.get('kpi3') == 'pass'])
        k3f = len([c for c in cases if c.get('kpi3') == 'fail'])
        k3t = k3p + k3f or 1

        ws_mc.cell(row=39, column=3).value  = k1p
        ws_mc.cell(row=39, column=5).value  = round(k1p / k1t * 100, 1)
        ws_mc.cell(row=41, column=3).value  = k1f
        ws_mc.cell(row=41, column=5).value  = round(k1f / k1t * 100, 1)
        ws_mc.cell(row=39, column=7).value  = k2p
        ws_mc.cell(row=39, column=9).value  = round(k2p / k2t * 100, 1)
        ws_mc.cell(row=41, column=7).value  = k2f
        ws_mc.cell(row=41, column=9).value  = round(k2f / k2t * 100, 1)
        ws_mc.cell(row=39, column=12).value = k3p
        ws_mc.cell(row=39, column=14).value = round(k3p / k3t * 100, 1)
        ws_mc.cell(row=41, column=12).value = k3f
        ws_mc.cell(row=41, column=14).value = round(k3f / k3t * 100, 1)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    fname = f"CM_Monthly_{month_iso}.xlsx"
    return send_file(out, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=fname)

if __name__ == '__main__':
    # สำหรับการนำไปรันในเครื่อง Local เครื่องจะดึง PORT จาก Environment หรือใช้ 5000 อัตโนมัติ
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)