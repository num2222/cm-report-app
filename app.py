"""
CM Daily Report — Flask Full-Stack
Deploy on Render.com with PostgreSQL
"""
import os, io, json
from flask import Flask, request, jsonify, send_file, render_template, session, redirect, url_for
from flask_sqlalchemy import SQLAlchemy
from functools import wraps
from datetime import datetime
from pathlib import Path
import random, string, time

app = Flask(__name__)

DATABASE_URL = os.environ.get('DATABASE_URL', 'sqlite:///cm_local.db')
if DATABASE_URL.startswith('postgres://'):
    DATABASE_URL = DATABASE_URL.replace('postgres://', 'postgresql://', 1)

app.config['SQLALCHEMY_DATABASE_URI'] = DATABASE_URL
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'cm-dev-key-change-in-prod')

# ── Login (รหัสผ่านเดียวร่วมกันทั้งทีม) ─────────────────
APP_PASSWORD = os.environ.get('APP_PASSWORD', 'changeme123')
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
# อายุ session: ล็อกอินค้างไว้ได้ 30 วัน จะได้ไม่ต้อง login ใหม่บ่อยๆ
app.config['PERMANENT_SESSION_LIFETIME'] = 60 * 60 * 24 * 30

db = SQLAlchemy(app)

def login_required(f):
    """decorator: บังคับให้ login ก่อนเข้าถึง route นี้
    หน้าเว็บ (HTML) → redirect ไปหน้า login
    API (/api/*) → ตอบกลับ 401 JSON แทน เพื่อให้ frontend จัดการต่อได้"""
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get('logged_in'):
            if request.path.startswith('/api/'):
                return jsonify(error='กรุณาเข้าสู่ระบบก่อนใช้งาน'), 401
            return redirect(url_for('login_page'))
        return f(*args, **kwargs)
    return wrapper

class Case(db.Model):
    __tablename__ = 'cases'
    id            = db.Column(db.String(20), primary_key=True)
    area          = db.Column(db.String(10), nullable=False, index=True)
    date          = db.Column(db.String(10), nullable=False, index=True)
    seq           = db.Column(db.String(20))
    job_no        = db.Column(db.String(100))
    sap_no        = db.Column(db.String(100), nullable=True)
    notify_time   = db.Column(db.String(20))
    kpi_val       = db.Column(db.String(50))
    response_time = db.Column(db.String(50))
    std           = db.Column(db.String(5))
    approve_time  = db.Column(db.String(20))
    arrive_time   = db.Column(db.String(20))
    start_time    = db.Column(db.String(20))
    close_time    = db.Column(db.String(20))
    close_date    = db.Column(db.String(10))   # วันที่ปิดงาน (กรณีปิดข้ามวัน)
    location      = db.Column(db.Text)
    problem       = db.Column(db.Text)
    solution      = db.Column(db.Text)
    reporter      = db.Column(db.String(200))
    receiver      = db.Column(db.String(200))
    technician    = db.Column(db.String(200))
    kpi1          = db.Column(db.String(10))
    kpi2          = db.Column(db.String(10))
    kpi3          = db.Column(db.String(10))
    # ── ปิด SAP ──────────────────────────────────────────
    sap_status    = db.Column(db.String(20), default='')
    sap_fail_reason = db.Column(db.Text)
    pending_reason  = db.Column(db.Text)
    # ── ยกเลิกใบงาน ───────────────────────────────────────
    cancelled     = db.Column(db.Boolean, default=False)
    created_at    = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at    = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    def to_dict(self):
        return {
            'id': self.id, 'area': self.area, 'date': self.date,
            'seq': self.seq, 'jobNo': self.job_no, 'sapNo': self.sap_no,
            'notifyTime': self.notify_time, 'kpiVal': self.kpi_val,
            'responseTime': self.response_time, 'std': self.std,
            'approveTime': self.approve_time, 'arriveTime': self.arrive_time,
            'startTime': self.start_time, 'closeTime': self.close_time,
            'closeDate': self.close_date or '',
            'location': self.location, 'problem': self.problem,
            'solution': self.solution, 'reporter': self.reporter,
            'receiver': self.receiver, 'technician': self.technician,
            'kpi1': self.kpi1, 'kpi2': self.kpi2, 'kpi3': self.kpi3,
            'sapStatus': self.sap_status or '',
            'sapFailReason': self.sap_fail_reason or '',
            'pendingReason': self.pending_reason or '',
            'cancelled': bool(self.cancelled),
        }

with app.app_context():
    db.create_all()
    # ── Auto-migration: เพิ่ม column ใหม่ในฐานข้อมูลเดิมโดยไม่ลบข้อมูล ──
    new_columns = [
        "ALTER TABLE cases ADD COLUMN IF NOT EXISTS sap_status VARCHAR(20) DEFAULT ''",
        "ALTER TABLE cases ADD COLUMN IF NOT EXISTS sap_fail_reason TEXT",
        "ALTER TABLE cases ADD COLUMN IF NOT EXISTS pending_reason TEXT",
        "ALTER TABLE cases ADD COLUMN IF NOT EXISTS cancelled BOOLEAN DEFAULT FALSE",
        "ALTER TABLE cases ADD COLUMN IF NOT EXISTS close_date VARCHAR(10)",
    ]
    try:
        with db.engine.connect() as conn:
            for sql in new_columns:
                try:
                    conn.execute(db.text(sql))
                except Exception:
                    pass  # column มีอยู่แล้ว
            conn.commit()
    except Exception:
        pass  # SQLite ไม่รองรับ IF NOT EXISTS — ไม่เป็นไร

def gen_id():
    return ''.join(random.choices(string.ascii_lowercase+string.digits, k=8))+str(int(time.time()))[-4:]

def fmt_be(d):
    if not d: return ''
    p = d.split('-')
    return f"{p[2]}/{p[1]}/{str(int(p[0])+543)[-2:]}"

def kpi_sym(v):
    """ใช้ Unicode เครื่องหมายถูก/กากบาทจริง แทนตัวอักษร P/O
    เพื่อให้แสดงผลเหมือนกันทุก cell ไม่ขึ้นกับ font ของ template เดิม"""
    return '✓' if v == 'pass' else ('✗' if v == 'fail' else '')

def _seq_num(c):
    """แปลงค่า seq (ลำดับเคสต่อวัน) เป็นตัวเลขสำหรับ sort"""
    try:
        return int(c.get('seq') or 0)
    except (ValueError, TypeError):
        return 999999

def _is_cancelled(c):
    """เช็คว่าเคสนี้ถูกยกเลิกใบงานหรือไม่"""
    return bool(c.get('cancelled'))

def _kpi2_export(c):
    """KPI2 สำหรับ export Excel:
    - ยกเลิกใบงาน → '' (ไม่แสดงผล KPI)
    - ไม่มีเวลาปิดงาน → fail เสมอ (ยังไม่เสร็จ)
    - มีเวลาปิดงาน → ใช้ค่า kpi2 จริง"""
    if _is_cancelled(c): return ''
    if not c.get('closeTime'): return 'fail'
    return c.get('kpi2', '')

def _kpi_export(c, key):
    """KPI1/KPI3 สำหรับ export: ถ้ายกเลิกใบงาน → '' (ไม่แสดง)"""
    if _is_cancelled(c): return ''
    return c.get(key, '')

STD_LABELS = {'A':'5-30 นาที','B':'1-3 ชม.','C':'3 ชม.-1 วัน','D':'1-7 วัน','E':'7-14 วัน','F':'1 เดือน'}

def _trunc(val, max_len):
    """ตัดความยาวสตริงไม่ให้เกิน max_len — ป้องกัน DataError จาก backend
    แม้ frontend parse ผิดพลาด ระบบจะไม่ crash แต่จะตัดข้อมูลส่วนเกินทิ้ง"""
    s = (val or '')
    return s[:max_len] if len(s) > max_len else s

def apply_dict(c, d):
    c.area         = _trunc(d.get('area',''), 10)
    c.date         = _trunc(d.get('date',''), 10)
    c.seq          = _trunc(d.get('seq',''), 20)
    c.job_no       = _trunc(d.get('jobNo',''), 100)
    sap = _trunc(d.get('sapNo','') or '', 100)
    c.sap_no       = sap or None
    c.notify_time  = _trunc(d.get('notifyTime',''), 20)
    c.kpi_val      = _trunc(d.get('kpiVal',''), 50)
    c.response_time= _trunc(d.get('responseTime',''), 50)
    c.std          = _trunc(d.get('std',''), 5)
    c.approve_time = _trunc(d.get('approveTime',''), 20)
    c.arrive_time  = _trunc(d.get('arriveTime',''), 20)
    c.start_time   = _trunc(d.get('startTime','') or d.get('arriveTime',''), 20)
    c.close_time   = _trunc(d.get('closeTime',''), 20)
    c.close_date   = _trunc(d.get('closeDate','') or '', 10) or None
    c.location     = d.get('location','')      # Text column — ไม่จำกัด
    c.problem      = d.get('problem','')        # Text column — ไม่จำกัด
    c.solution     = d.get('solution','')       # Text column — ไม่จำกัด
    c.reporter     = _trunc(d.get('reporter',''), 200)
    c.receiver     = _trunc(d.get('receiver',''), 200)
    c.technician   = _trunc(d.get('technician',''), 200)
    c.kpi1         = d.get('kpi1','')
    c.kpi2         = d.get('kpi2','')
    c.kpi3         = d.get('kpi3','')

    # ── ปิด SAP: auto-set สถานะตามว่ามี SAP No. หรือไม่ ──────
    # ถ้า frontend ส่ง sapStatus มาตรงๆ (กรณีแก้ไขในหน้า "ปิด SAP") ให้ใช้ตามนั้น
    incoming_status = d.get('sapStatus', None)
    if incoming_status is not None:
        c.sap_status = _trunc(incoming_status, 20)
    elif not sap:
        c.sap_status = ''  # ไม่มี SAP No. — ยังไม่เข้าสู่กระบวนการปิด SAP
    elif not c.sap_status:
        c.sap_status = 'pending'  # มี SAP No. แล้ว แต่ยังไม่เคยตั้งสถานะ → รอปิด SAP

    c.sap_fail_reason = d.get('sapFailReason', '') or ''
    c.pending_reason  = d.get('pendingReason', '') or ''
    c.cancelled       = bool(d.get('cancelled', False))
    return c

@app.route('/login', methods=['GET'])
def login_page():
    if session.get('logged_in'):
        return redirect(url_for('index'))
    return render_template('login.html')

@app.route('/api/login', methods=['POST'])
def do_login():
    d = request.get_json(silent=True) or {}
    password = d.get('password', '')
    if password == APP_PASSWORD:
        session.permanent = True
        session['logged_in'] = True
        return jsonify(ok=True)
    return jsonify(error='รหัสผ่านไม่ถูกต้อง'), 401

@app.route('/api/logout', methods=['POST'])
def do_logout():
    session.clear()
    return jsonify(ok=True)

@app.route('/')
@login_required
def index():
    return render_template('index.html')

@app.route('/ping')
def ping():
    try:
        db.session.execute(db.text('SELECT 1'))
        db_url = app.config['SQLALCHEMY_DATABASE_URI']
        db_type = 'postgresql' if 'postgresql' in db_url else 'sqlite (ยังไม่ได้เชื่อม PostgreSQL)'
        db_status = 'connected'
    except Exception as e:
        db_type = 'unknown'
        db_status = f'error: {str(e)}'
    return jsonify(ok=True, db=db_status, db_type=db_type)

@app.route('/debug/db-info')
def debug_db_info():
    """หน้าเช็คสถานะ database แบบละเอียด — เปิดผ่าน browser ได้เลย"""
    db_uri      = app.config['SQLALCHEMY_DATABASE_URI']
    is_postgres = db_uri.startswith('postgresql://')
    is_sqlite   = db_uri.startswith('sqlite://')

    info = {
        'db_type': 'PostgreSQL ✅' if is_postgres else
                   ('SQLite ⚠️ (ข้อมูลจะหายทุกครั้งที่ redeploy/restart)' if is_sqlite else 'Unknown'),
        'database_url_env_set': bool(os.environ.get('DATABASE_URL')),
        'connection_host': '',
        'total_cases': None,
        'sample_case_ids': [],
        'connection_ok': False,
        'error': None
    }

    try:
        if '@' in db_uri:
            info['connection_host'] = db_uri.split('@')[1].split('/')[0]
        else:
            info['connection_host'] = '(local sqlite file)'
    except Exception:
        pass

    try:
        count = Case.query.count()
        info['total_cases']    = count
        info['sample_case_ids']= [c.id for c in Case.query.limit(5).all()]
        info['connection_ok']  = True
    except Exception as e:
        info['error'] = str(e)

    return jsonify(info)

@app.route('/api/cases', methods=['GET'])
@login_required
def get_cases():
    area  = request.args.get('area')
    date  = request.args.get('date')
    month = request.args.get('month')
    q = Case.query
    if area and area != 'ALL': q = q.filter_by(area=area)
    if date:  q = q.filter_by(date=date)
    if month: q = q.filter(Case.date.like(f'{month}%'))
    cases = q.order_by(Case.date.desc(), Case.notify_time).all()
    return jsonify([c.to_dict() for c in cases])

@app.route('/api/cases', methods=['POST'])
@login_required
def add_case():
    d   = request.get_json()
    sap = (d.get('sapNo','') or '').strip()
    job = (d.get('jobNo','') or '').strip()
    if sap and Case.query.filter_by(sap_no=sap).first():
        return jsonify(error=f'SAP No. {sap} มีในฐานข้อมูลแล้ว'), 409
    if job and Case.query.filter_by(job_no=job).first():
        return jsonify(error=f'Job No. {job} มีในฐานข้อมูลแล้ว'), 409
    c   = apply_dict(Case(), d)
    c.id = d.get('id') or gen_id()
    db.session.add(c)
    db.session.commit()
    return jsonify(c.to_dict()), 201

@app.route('/api/cases/<cid>', methods=['PUT'])
@login_required
def update_case(cid):
    c = Case.query.get_or_404(cid)
    d = request.get_json()
    sap = (d.get('sapNo','') or '').strip()
    job = (d.get('jobNo','') or '').strip()
    if sap and sap != c.sap_no:
        dup = Case.query.filter_by(sap_no=sap).first()
        if dup and dup.id != cid:
            return jsonify(error=f'SAP No. {sap} มีในฐานข้อมูลแล้ว'), 409
    if job and job != c.job_no:
        dup = Case.query.filter_by(job_no=job).first()
        if dup and dup.id != cid:
            return jsonify(error=f'Job No. {job} มีในฐานข้อมูลแล้ว'), 409
    apply_dict(c, d)
    c.updated_at = datetime.utcnow()
    db.session.commit()
    return jsonify(c.to_dict())

@app.route('/api/cases/<cid>', methods=['DELETE'])
@login_required
def delete_case(cid):
    c = Case.query.get_or_404(cid)
    db.session.delete(c)
    db.session.commit()
    return jsonify(ok=True)

@app.route('/api/cases/<cid>/sap', methods=['PUT'])
@login_required
def update_sap(cid):
    """อัปเดตเฉพาะข้อมูล SAP (เลข SAP / สถานะ / เหตุผล) ใช้สำหรับ:
    - เพิ่มเลข SAP ทีหลัง (กรณีเคสที่ไม่มี SAP No. ตอนแรก)
    - เปลี่ยนสถานะ ปิด SAP / ปิด SAP ไม่สำเร็จ"""
    c = Case.query.get_or_404(cid)
    d = request.get_json()

    sap = (d.get('sapNo','') or '').strip()
    if sap and sap != c.sap_no:
        dup = Case.query.filter_by(sap_no=sap).first()
        if dup and dup.id != cid:
            return jsonify(error=f'SAP No. {sap} มีในฐานข้อมูลแล้ว'), 409
        c.sap_no = _trunc(sap, 100)
        # เพิ่ม SAP No. ทีหลัง → เริ่มสถานะเป็น "รอปิด SAP" ถ้ายังไม่เคยตั้งสถานะ
        if not c.sap_status:
            c.sap_status = 'pending'

    if 'sapStatus' in d:
        c.sap_status = _trunc(d.get('sapStatus','') or '', 20)
    if 'sapFailReason' in d:
        c.sap_fail_reason = d.get('sapFailReason','') or ''

    c.updated_at = datetime.utcnow()
    db.session.commit()
    return jsonify(c.to_dict())

def _write_rows(ws, cases, start_row=8):
    from openpyxl.styles import Font, Alignment
    # ── font หลักสำหรับข้อมูลทั่วไปทุก cell (TH SarabunPSK 12 เสมอ) ────────
    data_font     = Font(name='TH SarabunPSK', size=12)
    # ── font สำหรับ KPI result ✓/✗ (size 14 bold เพื่อให้เครื่องหมายชัด) ──
    kpi_font      = Font(name='TH SarabunPSK', size=14, bold=True, color='1E7E4A')
    kpi_font_fail = Font(name='TH SarabunPSK', size=14, bold=True, color='C0392B')
    center = Alignment(horizontal='center', vertical='center')
    wrap   = Alignment(wrap_text=True, vertical='center')
    left   = Alignment(vertical='center')

    def set_cell(ws, r, col, value, font=None, align=None):
        cell = ws.cell(row=r, column=col, value=value)
        # บังคับ font ทุก cell ไม่ให้ inherit จาก template เดิมที่อาจไม่สม่ำเสมอ
        cell.font      = font  or data_font
        cell.alignment = align or left
        return cell

    for i, c in enumerate(cases):
        r   = start_row + i
        std = c.get('std','')
        is_cancelled = c.get('cancelled', False)

        # ── reset row height ให้สม่ำเสมอทุก row ──────────────────────────────
        ws.row_dimensions[r].height = 20

        set_cell(ws, r,  1, c.get('seq') or (i+1),    align=center)
        set_cell(ws, r,  2, c.get('jobNo',''),         align=center)
        set_cell(ws, r,  3, c.get('sapNo',''),         align=center)
        set_cell(ws, r,  4, c.get('notifyTime',''),    align=center)
        # col E = พื้นที่ & บริเวณ (area + location รวมกัน)
        area_loc = ' '.join(filter(None, [c.get('area',''), c.get('location','')]))
        set_cell(ws, r,  5, area_loc,                  align=wrap)
        set_cell(ws, r,  6, c.get('problem',''),       align=wrap)
        set_cell(ws, r,  7, c.get('kpiVal',''),        align=center)
        set_cell(ws, r,  8, c.get('responseTime',''),  align=center)
        set_cell(ws, r,  9, c.get('approveTime',''),   align=center)
        set_cell(ws, r, 10, c.get('arriveTime',''),    align=center)

        if is_cancelled:
            set_cell(ws, r, 11, '', align=center)
        else:
            kpi1_val = c.get('kpi1')
            set_cell(ws, r, 11, kpi_sym(kpi1_val),
                     font=kpi_font if kpi1_val=='pass' else kpi_font_fail,
                     align=center)

        set_cell(ws, r, 12, c.get('solution',''),      align=wrap)
        set_cell(ws, r, 13, std,                       align=center)
        set_cell(ws, r, 14, STD_LABELS.get(std,''),    align=center)
        set_cell(ws, r, 15, c.get('closeTime',''),     align=center)

        # col Q = หมายเหตุ
        close_date = c.get('closeDate','')
        if is_cancelled:
            set_cell(ws, r, 16, '', align=center)
            set_cell(ws, r, 17, 'ยกเลิกใบงาน')
        else:
            kpi2_val = _kpi2_export(c)
            set_cell(ws, r, 16, kpi_sym(kpi2_val),
                     font=kpi_font if kpi2_val=='pass' else kpi_font_fail,
                     align=center)
            if close_date and close_date != c.get('date',''):
                set_cell(ws, r, 17, f'ปิดงานวันที่ {fmt_be(close_date)}')

@app.route('/api/export/daily', methods=['POST'])
@login_required
def export_daily():
    from openpyxl import load_workbook
    d        = request.get_json()
    date_iso = d.get('date','')
    cases_all= d.get('cases',[])
    tmpl = Path('Template_Daily_Report.xlsx')
    if not tmpl.exists():
        return jsonify(error='ไม่พบ Template_Daily_Report.xlsx'), 500
    wb = load_workbook(io.BytesIO(tmpl.read_bytes()))
    areas = ['MTB','Z2','FZ','SAT1']
    date_be = fmt_be(date_iso)
    area_cases = {}
    for area in areas:
        ws    = wb[area]
        cases = sorted([c for c in cases_all if c.get('area')==area], key=lambda c: _seq_num(c))
        area_cases[area] = cases
        done  = [c for c in cases if c.get('closeTime')]
        ws.cell(row=4,column=6).value  = date_be
        ws.cell(row=4,column=11).value = len(done)
        ws.cell(row=4,column=16).value = len(cases)-len(done)
        _write_rows(ws, cases, start_row=8)
    ws_cm = wb['Daily CM']
    all_c = [c for a in areas for c in area_cases[a]]
    all_d = [c for c in all_c if c.get('closeTime')]
    ws_cm.cell(row=7,column=5).value  = date_be
    ws_cm.cell(row=13,column=8).value = len(all_c)
    ws_cm.cell(row=14,column=8).value = len(all_d)
    ws_cm.cell(row=15,column=8).value = len(all_c)-len(all_d)
    for area,(r1,r2,r3) in {'MTB':(19,20,21),'Z2':(23,24,25),'FZ':(27,28,29),'SAT1':(31,32,33)}.items():
        ac=area_cases[area]; ad=[c for c in ac if c.get('closeTime')]
        ws_cm.cell(row=r1,column=10).value=len(ac)
        ws_cm.cell(row=r2,column=10).value=len(ad)
        ws_cm.cell(row=r3,column=10).value=len(ac)-len(ad)
    # KPI1: นับเฉพาะเคสที่ไม่ได้ยกเลิกใบงาน
    k1p=sum(1 for c in all_c if c.get('kpi1')=='pass' and not c.get('cancelled'))
    k1f=sum(1 for c in all_c if c.get('kpi1')=='fail' and not c.get('cancelled'))
    # KPI2: ไม่มีเวลาปิดงาน = ไม่ผ่าน (ยกเว้นเคสที่ยกเลิกใบงาน)
    k2p=sum(1 for c in all_c if _kpi2_export(c)=='pass' and not c.get('cancelled'))
    k2f=sum(1 for c in all_c if _kpi2_export(c)=='fail' and not c.get('cancelled'))
    # Daily CM: D38=KPI1 pass, D40=KPI1 fail, J38=KPI2 pass, J40=KPI2 fail
    # F38, L38, F40, L40 เป็น formula ใน template คำนวณ % อัตโนมัติ ไม่ต้องเขียน
    ws_cm.cell(row=38,column=4).value=k1p   # D38 KPI1 ผ่าน
    ws_cm.cell(row=40,column=4).value=k1f   # D40 KPI1 ไม่ผ่าน
    ws_cm.cell(row=38,column=10).value=k2p  # J38 KPI2 ผ่าน
    ws_cm.cell(row=40,column=10).value=k2f  # J40 KPI2 ไม่ผ่าน
    out = io.BytesIO(); wb.save(out); out.seek(0)
    return send_file(out, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'CM_Daily_{date_iso}.xlsx')

@app.route('/api/export/monthly', methods=['POST'])
@login_required
def export_monthly():
    from openpyxl import load_workbook
    d        = request.get_json()
    month    = d.get('month','')
    cases_all= d.get('cases',[])
    tmpl = Path('Template_Month_Report.xlsx')
    if not tmpl.exists():
        return jsonify(error='ไม่พบ Template_Month_Report.xlsx'), 500
    wb = load_workbook(io.BytesIO(tmpl.read_bytes()))
    yr,mo = month.split('-')
    month_be = f"{mo}/{int(yr)+543}"
    areas = ['MTB','Z2','FZ','SAT1']
    area_cases = {}

    def write_month_rows(ws, cases):
        from openpyxl.styles import Font, Alignment
        data_font     = Font(name='TH SarabunPSK', size=12)
        kpi_font_pass = Font(name='TH SarabunPSK', size=14, bold=True, color='1E7E4A')
        kpi_font_fail = Font(name='TH SarabunPSK', size=14, bold=True, color='C0392B')
        center = Alignment(horizontal='center', vertical='center')
        wrap   = Alignment(wrap_text=True, vertical='center')

        def set_cell(ws, r, col, value, font=None, align=None):
            cell = ws.cell(row=r, column=col, value=value)
            cell.font      = font  or data_font
            cell.alignment = align or Alignment(vertical='center')

        def set_kpi_cell(ws, r, col, val):
            f = kpi_font_pass if val == 'pass' else kpi_font_fail
            set_cell(ws, r, col, kpi_sym(val), font=f, align=center)

        def clear_kpi_cell(ws, r, col):
            set_cell(ws, r, col, '', align=center)

        for i,c in enumerate(cases):
            r=3+i; std=c.get('std','')
            is_cancelled = c.get('cancelled', False)
            close_date   = c.get('closeDate','')
            set_cell(ws, r,  1, c.get('seq') or (i+1))
            set_cell(ws, r,  2, c.get('jobNo',''))
            set_cell(ws, r,  3, c.get('sapNo',''))
            set_cell(ws, r,  4, c.get('notifyTime',''),   align=center)
            # col E = พื้นที่ & บริเวณ (area + location)
            area_loc = ' '.join(filter(None,[c.get('area',''), c.get('location','')]))
            set_cell(ws, r,  5, area_loc,                 align=wrap)
            set_cell(ws, r,  6, c.get('problem',''),      align=wrap)
            set_cell(ws, r,  7, c.get('kpiVal',''),       align=center)
            set_cell(ws, r,  8, c.get('responseTime',''), align=center)
            set_cell(ws, r,  9, c.get('approveTime',''),  align=center)
            set_cell(ws, r, 10, c.get('arriveTime',''),   align=center)
            set_cell(ws, r, 12, c.get('solution',''),     align=wrap)
            set_cell(ws, r, 13, std,                      align=center)
            set_cell(ws, r, 14, STD_LABELS.get(std,''))
            set_cell(ws, r, 15, c.get('closeTime',''),    align=center)
            if is_cancelled:
                clear_kpi_cell(ws, r, 11)
                clear_kpi_cell(ws, r, 16)
                clear_kpi_cell(ws, r, 17)
                set_cell(ws, r, 18, 'ยกเลิกใบงาน')  # col R = หมายเหตุ
            else:
                set_kpi_cell(ws, r, 11, c.get('kpi1'))
                set_kpi_cell(ws, r, 16, _kpi2_export(c))
                set_kpi_cell(ws, r, 17, c.get('kpi3'))
                if close_date and close_date != c.get('date',''):
                    set_cell(ws, r, 18, f'ปิดงานวันที่ {fmt_be(close_date)}')

    for area in areas:
        cases = sorted([c for c in cases_all if c.get('area')==area],
                       key=lambda c:(c.get('date',''), _seq_num(c)))
        area_cases[area] = cases
        write_month_rows(wb[area], cases)

    all_sorted = sorted([c for a in areas for c in area_cases[a]],
                        key=lambda c:(c.get('date',''), c.get('area',''), _seq_num(c)))
    write_month_rows(wb['ALL_ZONE'], all_sorted)

    ws_mc = wb['Month CM']
    ws_mc.cell(row=7,column=5).value = month_be
    all_c = all_sorted; all_d=[c for c in all_c if c.get('closeTime')]
    ws_mc.cell(row=13,column=8).value=len(all_c)
    ws_mc.cell(row=14,column=8).value=len(all_d)
    ws_mc.cell(row=15,column=8).value=len(all_c)-len(all_d)
    for area,(r1,r2,r3) in {'MTB':(18,19,20),'Z2':(21,22,23),'FZ':(24,25,26),'SAT1':(27,28,29)}.items():
        ac=area_cases[area]; ad=[c for c in ac if c.get('closeTime')]
        ws_mc.cell(row=r1,column=13).value=len(ac)
        ws_mc.cell(row=r2,column=13).value=len(ad)
        ws_mc.cell(row=r3,column=13).value=len(ac)-len(ad)
    k1p=sum(1 for c in all_c if c.get('kpi1')=='pass' and not c.get('cancelled'))
    k1f=sum(1 for c in all_c if c.get('kpi1')=='fail' and not c.get('cancelled'))
    k2p=sum(1 for c in all_c if _kpi2_export(c)=='pass' and not c.get('cancelled'))
    k2f=sum(1 for c in all_c if _kpi2_export(c)=='fail' and not c.get('cancelled'))
    k3p=sum(1 for c in all_c if c.get('kpi3')=='pass' and not c.get('cancelled'))
    k3f=sum(1 for c in all_c if c.get('kpi3')=='fail' and not c.get('cancelled'))
    k1t=(k1p+k1f) or 1; k2t=(k2p+k2f) or 1; k3t=(k3p+k3f) or 1
    # Month CM cell addresses ตาม Template จริง:
    # KPI1: C39=pass count, E39=pass%, C41=fail count, E41=fail%
    # KPI2: G39=pass count, I39=pass%, G41=fail count, I41=fail%
    # KPI3: L39=pass count, N39=pass%, L41=fail count, N41=fail%
    ws_mc.cell(row=39,column=3).value=k1p;   ws_mc.cell(row=39,column=5).value=round(k1p/k1t*100,1)   # C39, E39
    ws_mc.cell(row=41,column=3).value=k1f;   ws_mc.cell(row=41,column=5).value=round(k1f/k1t*100,1)   # C41, E41
    ws_mc.cell(row=39,column=7).value=k2p;   ws_mc.cell(row=39,column=9).value=round(k2p/k2t*100,1)   # G39, I39
    ws_mc.cell(row=41,column=7).value=k2f;   ws_mc.cell(row=41,column=9).value=round(k2f/k2t*100,1)   # G41, I41
    ws_mc.cell(row=39,column=12).value=k3p;  ws_mc.cell(row=39,column=14).value=round(k3p/k3t*100,1)  # L39, N39
    ws_mc.cell(row=41,column=12).value=k3f;  ws_mc.cell(row=41,column=14).value=round(k3f/k3t*100,1)  # L41, N41
    out = io.BytesIO(); wb.save(out); out.seek(0)
    return send_file(out, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'CM_Monthly_{month}.xlsx')

@app.route('/api/export/sap', methods=['POST'])
@login_required
def export_sap():
    """Export หน้าปิด SAP เป็น Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    d = request.get_json()
    cases = d.get('cases', [])
    filters = d.get('filters', {})

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'ปิด SAP'

    # Header
    ws.merge_cells('A1:J1')
    ws['A1'] = 'รายงานปิด SAP — ระบบไฟฟ้าแรงดันต่ำ (LV System)'
    ws['A1'].font = Font(bold=True, size=13)
    ws['A1'].alignment = Alignment(horizontal='center')

    filter_desc = []
    if filters.get('area') and filters['area'] != 'ALL':
        filter_desc.append(f"พื้นที่: {filters['area']}")
    if filters.get('from'): filter_desc.append(f"จาก: {filters['from']}")
    if filters.get('to'):   filter_desc.append(f"ถึง: {filters['to']}")
    if filters.get('status') and filters['status'] != 'ALL':
        status_map = {'pending':'รอปิด SAP','closed':'ปิด SAP','failed':'ปิด SAP ไม่สำเร็จ'}
        filter_desc.append(f"สถานะ: {status_map.get(filters['status'], filters['status'])}")
    ws.merge_cells('A2:J2')
    ws['A2'] = '  '.join(filter_desc) if filter_desc else 'ทุกพื้นที่ / ทุกสถานะ'
    ws['A2'].alignment = Alignment(horizontal='center')
    ws['A2'].font = Font(size=11, color='555555')

    # Column headers
    headers = ['พื้นที่','วันที่','SAP No.','Job No.','บริเวณ','ปัญหา','KPI3','สถานะ SAP','เหตุผล (ถ้าไม่สำเร็จ)']
    col_widths = [8, 10, 12, 10, 22, 28, 8, 14, 30]
    hdr_fill = PatternFill('solid', fgColor='1A3A5C')
    hdr_font = Font(color='FFFFFF', bold=True, size=11)
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=3, column=ci, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin
        ws.column_dimensions[get_column_letter(ci)].width = w

    status_map_th = {'pending':'รอปิด SAP', 'closed':'ปิด SAP', 'failed':'ปิด SAP ไม่สำเร็จ'}
    kpi3_map = {'pass':'ผ่าน', 'fail':'ไม่ผ่าน', '':'N/A', None:'N/A'}

    for ri, c in enumerate(cases, 4):
        st = c.get('sapStatus') or 'pending'
        row_data = [
            c.get('area',''), fmt_be(c.get('date','')),
            c.get('sapNo',''), c.get('jobNo',''),
            c.get('location',''), c.get('problem',''),
            kpi3_map.get(c.get('kpi3'),'N/A'),
            status_map_th.get(st, st),
            c.get('sapFailReason','') if st == 'failed' else ''
        ]
        fill_color = {'closed':'D4EDDA', 'failed':'F8D7DA', 'pending':'FFF3CD'}.get(st, 'FFFFFF')
        row_fill = PatternFill('solid', fgColor=fill_color)
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = thin
            cell.fill = row_fill
            cell.alignment = Alignment(vertical='center', wrap_text=True)

    ws.row_dimensions[3].height = 20

    # Summary row
    summary_row = len(cases) + 4
    ws.cell(row=summary_row, column=1, value=f'รวม {len(cases)} เคส')
    ws.cell(row=summary_row, column=1).font = Font(bold=True)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(out, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='CM_SAP_Status.xlsx')

# ── BACKUP / IMPORT ────────────────────────────────────────────────────
@app.route('/api/backup', methods=['GET'])
@login_required
def backup_db():
    """Export ข้อมูลทั้งหมดเป็น JSON file สำหรับ backup
    เรียกผ่าน GET /api/backup → ดาวน์โหลดไฟล์ cm_backup_YYYYMMDD.json"""
    cases = Case.query.order_by(Case.date, Case.area, Case.seq).all()
    data = {
        'version': '1.0',
        'exported_at': datetime.utcnow().isoformat(),
        'total': len(cases),
        'cases': [c.to_dict() for c in cases]
    }
    json_bytes = json.dumps(data, ensure_ascii=False, indent=2).encode('utf-8')
    buf = io.BytesIO(json_bytes)
    buf.seek(0)
    from datetime import date
    filename = f"cm_backup_{date.today().strftime('%Y%m%d')}.json"
    return send_file(buf, mimetype='application/json',
                     as_attachment=True, download_name=filename)

@app.route('/api/import', methods=['POST'])
@login_required
def import_db():
    """Import ข้อมูลจากไฟล์ JSON backup กลับเข้า database
    Body: multipart/form-data
      - file: ไฟล์ .json จาก /api/backup
      - mode: 'merge' (default) = เพิ่มเคสที่ยังไม่มีใน DB
              'replace' = ลบทั้งหมดแล้ว import ใหม่ (ต้องส่ง confirm=true ด้วย)"""

    # อ่าน mode/confirm จาก form data ก่อนเสมอ (multipart)
    ct = request.content_type or ''
    if 'multipart' in ct or 'form' in ct:
        mode    = request.form.get('mode', 'merge')
        confirm = request.form.get('confirm', '')
    else:
        d       = request.get_json(silent=True) or {}
        mode    = d.get('mode', 'merge')
        confirm = str(d.get('confirm', ''))

    # รับ file จาก multipart upload
    if 'file' not in request.files:
        return jsonify(error='ไม่พบ file ใน request — ส่งเป็น multipart/form-data field "file"'), 400

    f = request.files['file']
    try:
        raw = f.read().decode('utf-8')
        payload = json.loads(raw)
    except Exception as e:
        return jsonify(error=f'ไฟล์ไม่ถูกต้อง: {e}'), 400

    cases_data = payload.get('cases', [])
    if not cases_data:
        return jsonify(error='ไม่พบข้อมูล cases ในไฟล์'), 400

    if mode == 'replace':
        if str(confirm).lower() not in ('true', '1', 'yes'):
            return jsonify(error='mode=replace ต้องส่ง confirm=true เพื่อยืนยันการลบข้อมูลเดิมทั้งหมด'), 400
        Case.query.delete()
        db.session.flush()

    imported = skipped = errors = 0
    for d in cases_data:
        try:
            existing = Case.query.get(d.get('id'))
            if existing and mode == 'merge':
                skipped += 1
                continue
            if existing:
                apply_dict(existing, d)
                existing.updated_at = datetime.utcnow()
            else:
                c = apply_dict(Case(), d)
                c.id = d.get('id') or gen_id()
                db.session.add(c)
            imported += 1
        except Exception as e:
            app.logger.error(f"Import error for case {d.get('id')}: {e}")
            errors += 1

    db.session.commit()
    return jsonify(
        ok=True,
        mode=mode,
        imported=imported,
        skipped=skipped,
        errors=errors,
        total_in_db=Case.query.count()
    )


def handle_500(e):
    """ทุก error ที่ไม่ถูกจับจะ return JSON แทน HTML error page เสมอ"""
    db.session.rollback()
    app.logger.error(f"Unhandled 500 error: {e}")
    return jsonify(error=f'เกิดข้อผิดพลาดที่ server: {str(e)}'), 500

@app.errorhandler(404)
def handle_404(e):
    return jsonify(error='ไม่พบ endpoint ที่ร้องขอ'), 404

@app.errorhandler(Exception)
def handle_exception(e):
    """จับ exception ทุกชนิดที่ไม่มี handler เฉพาะ — กัน DataError, IntegrityError ฯลฯ"""
    db.session.rollback()
    app.logger.error(f"Unhandled exception: {type(e).__name__}: {e}")
    code = getattr(e, 'code', 500)
    if not isinstance(code, int):
        code = 500
    return jsonify(error=f'เกิดข้อผิดพลาด: {str(e)}'), code

if __name__ == '__main__':
    app.run(debug=True, port=5000)
